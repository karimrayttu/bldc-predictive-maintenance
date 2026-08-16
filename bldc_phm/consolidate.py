"""Consolidate all recorded runs into a single, clean, professional Excel workbook."""
import os
import re
import glob
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# grouping identity
IDENTITY = ["fixture_state", "condition", "severity", "config"]
_RPM_TOKEN = re.compile(r"_\d+rpm$")

# human-readable headers (raw column -> friendly name w/ unit)
FRIENDLY = {
    "session_id": "Run", "speed_setpoint_rpm": "Setpoint (rpm)", "severity": "Severity",
    "iso_time": "Timestamp (UTC)", "t_ms": "t (ms)",
    "rpm": "Speed (rpm)", "ripple_permil": "Ripple (per-mille)",
    "min_us": "Hall min (us)", "max_us": "Hall max (us)",
    "hall": "Hall state", "illegal": "Illegal Hall (cum)",
    "edges": "Hall edges (cum)", "step": "Cmd source", "coast_ms": "Coast (ms)",
    "period_std_permil": "Period std (per-mille)", "mech1x_permil": "Mech 1x (per-mille)",
    "elec_permil": "Elec order (per-mille)",
    "accel_rms_mg": "Vib RMS (mg)", "accel_pk_mg": "Vib peak (mg)",
    "accel_x_mg": "Grav X (mg)", "accel_y_mg": "Grav Y (mg)", "accel_z_mg": "Grav Z (mg)",
    "temp_mv": "NTC raw (mV)", "i_dc_mv": "Idc sensor (mV)",
    "adc_vref_raw": "VREFINT (raw)", "adc_pc0_mv": "PC0 diag (mV)",
    "temp_motor_c": "Motor temp (C)", "temp_ambient_c": "Ambient temp (C)",
    "i_u_a": "Iu (A)", "i_v_a": "Iv (A)", "i_w_a": "Iw (A)", "i_dc_a": "Idc (A)",
    "modbus_rpm": "Modbus rpm", "modbus_fault": "Modbus fault",
    # health + envelope; these were absent from the deliverable entirely, which meant a
    # reader could not tell a live reading from a frozen one, and the two columns schema.py
    # calls "THE columns to train on" never reached the workbook at all.
    "sensor_status": "Sensor status (bitmask)",
    "i_dc_peak_mv": "Idc peak (mV)", "i_dc_min_mv": "Idc min (mV)",
    "i_dc_peak_a": "Idc peak (A)", "i_dc_min_a": "Idc min (A)",
    "i_dc_a_signed": "Idc signed (A)", "i_dc_zero_mv": "Idc zero (mV)",
    "temp_motor_f": "Motor temp (F)",
    "valid_for_classifier": "Classifier-valid",
    "condition": "Condition", "severity": "Severity",
}
# preferred left-to-right order on a data sheet (anything else appended after)
DATA_ORDER = ["session_id", "condition", "severity", "speed_setpoint_rpm",
              "valid_for_classifier", "iso_time", "t_ms", "rpm",
              "ripple_permil", "period_std_permil", "mech1x_permil", "elec_permil",
              "min_us", "max_us", "hall", "illegal", "edges",
              "step", "coast_ms",
              "accel_rms_mg", "accel_pk_mg", "accel_x_mg", "accel_y_mg", "accel_z_mg",
              "temp_mv", "temp_motor_c", "temp_motor_f",
              "i_dc_mv", "i_dc_a", "i_dc_a_signed", "i_dc_zero_mv",
              "i_dc_peak_mv", "i_dc_min_mv", "i_dc_peak_a", "i_dc_min_a",
              "sensor_status", "adc_vref_raw", "adc_pc0_mv",
              "temp_ambient_c", "i_u_a", "i_v_a", "i_w_a",
              "modbus_rpm", "modbus_fault"]
# columns that are constant per group -> shown in the caption, not as table columns
# Shown in the caption line AND kept as columns. They used to be caption-only, which
# is why `healthy` came out 22 columns wide and every other sheet 23; the sheets could
# not be stacked without realigning them by hand.
CAPTION_COLS = {"fixture_state", "config", "label"}
# core signal columns always kept (even if constant); other columns are dropped when
# they are empty OR constant across the whole sheet (removes noise like step=9, coast=0).
# Never pruned, even when constant on a sheet. A column that vanishes from one sheet
# because it happened not to vary there is what makes sheets non-unionable, and a
# health column is MOST informative exactly when it is constant.
ALWAYS_KEEP = {"session_id", "condition", "severity", "speed_setpoint_rpm",
               "valid_for_classifier", "iso_time", "t_ms", "rpm",
               "ripple_permil", "period_std_permil", "mech1x_permil", "elec_permil",
               "min_us", "max_us", "hall", "illegal", "edges",
               "accel_rms_mg", "accel_pk_mg", "accel_x_mg", "accel_y_mg", "accel_z_mg",
               "temp_mv", "temp_motor_c", "temp_motor_f",
               "i_dc_mv", "i_dc_a", "i_dc_a_signed", "i_dc_zero_mv",
               "i_dc_peak_mv", "i_dc_min_mv", "i_dc_peak_a", "i_dc_min_a",
               "sensor_status"}
# per-column Excel number formats (friendly-name -> format)
NUMFMT = {"Ripple (per-mille)": "0", "Speed (rpm)": "0", "Setpoint (rpm)": "0",
          "Speed err (%)": "0.00", "Speed mean (rpm)": "0.0",
          "Ripple mean (per-mille)": "0.0", "Duration (s)": "0.0",
          "Vib RMS (mg)": "0", "Vib peak (mg)": "0",
          "NTC raw (mV)": "0", "Idc sensor (mV)": "0", "Idc (A)": "0.000",
          "Motor temp (C)": "0.0", "Motor temp (F)": "0.0", "Ambient temp (C)": "0.0",
          "Idc peak (A)": "0.000", "Idc min (A)": "0.000", "Idc signed (A)": "0.000",
          "Idc zero (mV)": "0.0", "Idc peak (mV)": "0", "Idc min (mV)": "0",
          "Sensor status (bitmask)": "0"}

# Excel hard limit. 150 runs x 5 min x 10 Hz is ~450k rows today, but a repeated
# campaign crosses this, and openpyxl fails late and unhelpfully when it does.
EXCEL_MAX_ROWS = 1_048_576

# styling
_HDR_FILL = PatternFill("solid", fgColor="1F3A5F")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_CAP_FILL = PatternFill("solid", fgColor="E8EEF5")
_CAP_FONT = Font(bold=True, size=12, color="1F3A5F")
_THIN = Side(style="thin", color="D0D7DE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _config(label):
    return _RPM_TOKEN.sub("", str(label))


def _sheet_name(used, fixture, config, severity):
    fix = "M" if str(fixture).startswith("mounted") else "C"
    sev = "" if str(severity) in ("none", "", "nan") else f"_{severity}"
    base = re.sub(r"[^0-9A-Za-z_]", "", f"{fix}_{config}{sev}")[:31] or "sheet"
    name, n = base, 1
    while name.lower() in used:
        sfx = f"~{n}"; name = base[:31 - len(sfx)] + sfx; n += 1
    used.add(name.lower())
    return name


def _run_stats(df):
    span = (df["t_ms"].iloc[-1] - df["t_ms"].iloc[0]) / 1000.0 if len(df) > 1 else 0.0
    illegal = int(df["illegal"].iloc[-1] - df["illegal"].iloc[0]) if "illegal" in df else 0
    setp = df["speed_setpoint_rpm"].iloc[0]
    rpm_mean = df["rpm"].mean()
    return {
        "Fixture": df["fixture_state"].iloc[0], "Condition": df["condition"].iloc[0],
        "Severity": df["severity"].iloc[0], "Config": _config(df["label"].iloc[0]),
        "Run": df["session_id"].iloc[0], "Setpoint (rpm)": setp,
        "Speed mean (rpm)": round(rpm_mean, 1), "Speed min": int(df["rpm"].min()),
        "Speed max": int(df["rpm"].max()),
        "Speed err (%)": round((rpm_mean - setp) / setp * 100, 2) if setp else None,
        "Ripple mean (per-mille)": round(df["ripple_permil"].mean(), 1),
        "Illegal events": illegal, "Samples": len(df), "Duration (s)": round(span, 1),
    }


def _speeds_with_counts(dfs):
    speeds = sorted(int(d["speed_setpoint_rpm"].iloc[0]) for d in dfs)
    from collections import Counter
    c = Counter(speeds)
    return ", ".join(f"{s}x{c[s]}" if c[s] > 1 else f"{s}" for s in sorted(c))


def _autofit(ws, header_row, max_width=26):
    for col in ws.iter_cols(min_row=header_row, max_row=ws.max_row):
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 8), max_width)


def _style_header(ws, row, ncols):
    for j in range(1, ncols + 1):
        c = ws.cell(row=row, column=j)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER


def _apply_numfmt(ws, header_row, headers):
    for j, h in enumerate(headers, start=1):
        fmt = NUMFMT.get(h)
        if not fmt:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=r, column=j).number_format = fmt


RD_MODE = "rd"   # R&D / debugging runs live under data/rd, kept OUT of the official dataset


def _mode_of(path, data_dir):
    rel = os.path.relpath(path, data_dir)
    return rel.split(os.sep)[0]


def verify_run(folder):
    """Confirm a saved run holds REAL sensor data (not fabricated/stuck/simulated).
    Returns (verdict, issues). verdict: VERIFIED | SIMULATED | SUSPECT."""
    import json
    issues = []
    meta = {}
    mp = os.path.join(folder, "meta.json")
    if os.path.exists(mp):
        try:
            meta = json.load(open(mp))
        except Exception:
            meta = {}
    src = meta.get("data_source") or meta.get("source") or "?"
    try:
        df = pd.read_csv(os.path.join(folder, "data.csv"))
    except Exception:
        return "SUSPECT", ["unreadable data.csv"]
    if len(df) < 10:
        issues.append("too few rows")
    if "hall" in df and df["hall"].nunique() < 4:
        issues.append("Hall stuck (<4 distinct states)")
    if {"t_ms", "edges", "rpm"} <= set(df.columns) and len(df) > 20:
        dt = (df["t_ms"].iloc[-1] - df["t_ms"].iloc[0]) / 1000.0
        de = df["edges"].iloc[-1] - df["edges"].iloc[0]
        mean_rpm = df["rpm"].mean()
        if dt > 0 and mean_rpm > 0:
            meas, exp = de / dt, mean_rpm * 24 / 60       # 24 Hall edges / mech rev
            if abs(meas - exp) / max(exp, 1) > 0.25:
                issues.append(f"rpm↔edges mismatch ({meas:.0f}/s vs {exp:.0f}/s)")
    if "ripple_permil" in df and len(df) and not (0 < df["ripple_permil"].median() < 800):
        issues.append("ripple out of physical band")
    if "rpm" in df and len(df) > 10 and float(df["rpm"].std()) == 0.0:
        issues.append("rpm frozen (no jitter; not a live sensor)")
    # Runs recorded before the field existed leave src empty and are judged on
    # their data alone. "simulator" appears only in datasets predating the
    # simulator's removal; anything else non-empty means the app could not
    # confirm the board was the source.
    if src and src not in ("serial", "simulator"):
        issues.append(f"data source not confirmed as the board ({src})")
    verdict = ("SIMULATED" if src == "simulator"
               else "VERIFIED" if not issues else "SUSPECT")
    return verdict, issues


def verify_dataset(data_dir, rd=False):
    """Verify every saved run (official, or rd=True for R&D). Returns list of
    (run_folder_name, verdict, issues)."""
    folders = [os.path.dirname(f)
               for f in glob.glob(os.path.join(data_dir, "*", "sessions", "**", "data.csv"),
                                  recursive=True)
               if (_mode_of(f, data_dir) == RD_MODE) == rd]
    out = []
    for fd in sorted(folders):
        v, iss = verify_run(fd)
        out.append((os.path.basename(fd), v, iss))
    return out


def consolidate(data_dir, out_path, rd=False):
    """Build the consolidated workbook. rd=False -> official datasets only (excludes data/rd);
    rd=True -> ONLY the R&D/debug runs. Returns a small report dict."""
    files = sorted(glob.glob(os.path.join(data_dir, "*", "sessions", "**", "data.csv"),
                             recursive=True))
    files = [f for f in files
             if (_mode_of(f, data_dir) == RD_MODE) == rd]      # official vs R&D split
    if not files:
        raise FileNotFoundError("No R&D run data found yet." if rd
                                else "No official run data found yet (no sessions/**/data.csv).")

    frames, summary_rows = [], []
    for f in files:
        df = pd.read_csv(f)
        if df.empty or "label" not in df.columns:
            continue
        df = df.copy(); df["config"] = _config(df["label"].iloc[0])
        frames.append((tuple(df[k].iloc[0] for k in IDENTITY), df))
        summary_rows.append(_run_stats(df))

    if not frames:
        raise FileNotFoundError("Run files exist but contain no rows yet.")

    groups = {}
    for ident, df in frames:
        groups.setdefault(ident, []).append(df)

    used, index_rows, sheet_for = set(), [], {}
    for ident, dfs in groups.items():
        fixture, condition, severity, config = ident
        name = _sheet_name(used, fixture, config, severity)
        sheet_for[ident] = name
        index_rows.append({
            "Sheet": name, "Fixture": fixture, "Condition": condition,
            "Severity": severity, "Config": config,
            "Classifier-valid": dfs[0]["valid_for_classifier"].iloc[0],
            "Runs": len(dfs), "Speeds (rpm)": _speeds_with_counts(dfs),
            "Samples": sum(len(d) for d in dfs),
        })

    index_df = pd.DataFrame(index_rows)
    summary_df = (pd.DataFrame(summary_rows)
                  .sort_values(["Fixture", "Condition", "Severity", "Config",
                                "Setpoint (rpm)", "Run"]))

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    # ATOMIC (2026-07-30): the app is killed regularly by an external UI-automation
    # collision, and one such crash mid-write corrupted MOUNTED.xlsx (BadZipFile).
    # Write to a temp file and os.replace so the workbook on disk is always whole.
    _tmp_path = out_path + ".tmp.xlsx"
    with pd.ExcelWriter(_tmp_path, engine="openpyxl") as xl:
        index_df.to_excel(xl, sheet_name="Index", index=False)
        summary_df.to_excel(xl, sheet_name="Run_Summary", index=False)

        for ident, dfs in groups.items():
            fixture, condition, severity, config = ident
            block = pd.concat(dfs, ignore_index=True)
            # keep only columns that vary within the sheet AND are not all-empty
            cols = [c for c in DATA_ORDER if c in block.columns and c not in CAPTION_COLS]
            # keep non-empty columns; drop columns that are constant across the sheet
            # unless they're core signal columns
            kept = []
            for c in cols:
                nonempty = block[c].notna().any() and block[c].astype(str).str.len().gt(0).any()
                if not nonempty:
                    continue
                if c not in ALWAYS_KEEP and block[c].nunique(dropna=False) <= 1:
                    continue
                kept.append(c)
            cols = kept
            block = block[cols].sort_values(
                [c for c in ("speed_setpoint_rpm", "session_id", "t_ms") if c in cols])
            block = block.rename(columns={c: FRIENDLY.get(c, c) for c in cols})
            block.to_excel(xl, sheet_name=sheet_for[ident], index=False, startrow=2)

        # styling pass
        book = xl.book
        # Index + Run_Summary
        for sname, df in (("Index", index_df), ("Run_Summary", summary_df)):
            ws = book[sname]
            _style_header(ws, 1, df.shape[1])
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            _apply_numfmt(ws, 1, list(df.columns))
            _autofit(ws, 1)
        # group data sheets: caption (row1) + header (row3)
        for ident, dfs in groups.items():
            fixture, condition, severity, config = ident
            ws = book[sheet_for[ident]]
            ncols = ws.max_column
            # caption from known identity parts (don't split tokens like "no_load")
            extras = config[len(str(condition)) + 1:] if config.startswith(str(condition) + "_") else config
            parts = [str(condition)]
            if str(severity) not in ("none", ""):
                parts.append(str(severity))
            if extras:
                parts.append(extras)
            caption = (" / ".join(parts) + f"   [{fixture}]"
                       f"   -   {len(dfs)} run(s), {_speeds_with_counts(dfs)} rpm")
            ws.cell(row=1, column=1, value=caption)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            cap = ws.cell(row=1, column=1)
            cap.fill = _CAP_FILL; cap.font = _CAP_FONT
            cap.alignment = Alignment(horizontal="left", vertical="center")
            headers = [ws.cell(row=3, column=j).value for j in range(1, ncols + 1)]
            _style_header(ws, 3, ncols)
            ws.freeze_panes = "C4"           # lock caption+header and Run/Setpoint cols
            ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{ws.max_row}"
            _apply_numfmt(ws, 3, headers)
            _autofit(ws, 3)

    os.replace(_tmp_path, out_path)
    return {"out_path": out_path, "runs": len(frames), "groups": len(groups),
            "sheets": ["Index", "Run_Summary"] + list(sheet_for.values())}


# ============================================================================
#  Simple per-fixture workbook: ONE sheet per category, rows grouped by rpm.
#  Official runs only (R&D + coast-downs excluded). This is the main deliverable.
# ============================================================================
import json as _json

# The column set for the per-condition sheets. This is the artefact a human opens and
# the one a model gets trained from, so it carries EVERYTHING needed to judge a row:
# the health bitmask, the current envelope (peak/min, schema.py calls these "THE columns
# to train on"), the unclamped signed current, and the learned zero that makes amps
# reproducible from raw mV. All of those were previously absent from the workbook.
FIXTURE_ORDER = ["session_id", "condition", "severity", "speed_setpoint_rpm",
                 "valid_for_classifier", "iso_time", "t_ms",
                 "rpm", "ripple_permil", "period_std_permil", "mech1x_permil", "elec_permil",
                 "min_us", "max_us", "hall", "illegal", "edges", "step", "coast_ms",
                 "accel_rms_mg", "accel_pk_mg", "accel_x_mg", "accel_y_mg", "accel_z_mg",
                 "temp_mv", "temp_motor_c", "temp_motor_f",
                 "i_dc_mv", "i_dc_a", "i_dc_a_signed", "i_dc_zero_mv",
                 "i_dc_peak_mv", "i_dc_min_mv", "i_dc_peak_a", "i_dc_min_a",
                 "sensor_status", "adc_vref_raw",
                 "temp_ambient_c", "i_u_a", "i_v_a", "i_w_a", "modbus_rpm", "modbus_fault"]


def _official_folders(data_dir, subroot):
    """Folders of OFFICIAL full-cycle runs under data/<subroot>/sessions (no R&D, no coast-downs)."""
    out = []
    base = os.path.join(data_dir, subroot, "sessions")
    for m in glob.glob(os.path.join(base, "**", "meta.json"), recursive=True):
        try:
            meta = _json.load(open(m))
        except Exception:
            continue
        if meta.get("run_kind") == "rd" or meta.get("test_kind") == "coastdown":
            continue
        if not meta.get("closed_utc"):
            continue          # crash fragment / still recording; never a deliverable
        if str(meta.get("valid_for_classifier")).lower() == "false":
            continue          # flagged runs stay on disk but out of the workbook
        out.append(os.path.dirname(m))
    return out


def _sanitize_sheet(name):
    return re.sub(r"[^0-9A-Za-z_]", "_", str(name))[:31] or "sheet"


def consolidate_fixture(data_dir, subroot, out_path):
    """Build ONE workbook for a fixture: an Index + one sheet per category (condition),
    rows stacked and sorted by rpm. Official runs only. Returns a small report."""
    folders = _official_folders(data_dir, subroot)
    by_cond = {}
    for fd in sorted(folders):
        try:
            df = pd.read_csv(os.path.join(fd, "data.csv"))
        except Exception:
            continue
        if df.empty or "condition" not in df.columns:
            continue
        by_cond.setdefault(str(df["condition"].iloc[0]), []).append(df)
    if not by_cond:
        raise FileNotFoundError(f"No official runs yet for {subroot}.")

    from collections import Counter
    index_rows = []
    blocks = {}
    for cond in sorted(by_cond):
        dfs = by_cond[cond]
        blocks[cond] = pd.concat(dfs, ignore_index=True)
        spc = Counter()
        for d in dfs:
            try:
                spc[int(float(d["speed_setpoint_rpm"].iloc[0]))] += 1
            except (TypeError, ValueError, KeyError):
                pass
        for sp in sorted(spc):
            index_rows.append({"Category": cond, "Speed (rpm)": sp, "Runs": spc[sp]})
    index_df = pd.DataFrame(index_rows)

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    sheet_for = {}
    overflow = []
    # ATOMIC (2026-07-30): the app is killed regularly by an external UI-automation
    # collision, and one such crash mid-write corrupted MOUNTED.xlsx (BadZipFile).
    # Write to a temp file and os.replace so the workbook on disk is always whole.
    _tmp_path = out_path + ".tmp.xlsx"
    with pd.ExcelWriter(_tmp_path, engine="openpyxl") as xl:
        index_df.to_excel(xl, sheet_name="Index", index=False)
        for cond in sorted(by_cond):
            block = blocks[cond].copy()
            block["_sp"] = pd.to_numeric(block.get("speed_setpoint_rpm"), errors="coerce")
            # Identical columns on every sheet. Pruning per-sheet (dropping `severity`
            # where it was all "none", dropping any all-empty column) is what made the
            # healthy sheet one column narrower than the others; the sheets then could
            # not be concatenated without realigning them by hand, which is the first
            # thing anyone does before training.
            cols = [c for c in FIXTURE_ORDER if c in block.columns]
            out = block.sort_values(["_sp"] + [c for c in ("session_id", "t_ms") if c in block.columns])
            out = out[cols].rename(columns={c: FRIENDLY.get(c, c) for c in cols})
            name = _sanitize_sheet(cond); sheet_for[cond] = name
            if len(out) > EXCEL_MAX_ROWS - 8:
                csv_path = os.path.splitext(out_path)[0] + f"_{_sanitize_sheet(cond)}.csv"
                out.to_csv(csv_path, index=False)
                overflow.append((cond, len(out), csv_path))
                out = out.head(EXCEL_MAX_ROWS - 8)
            out.to_excel(xl, sheet_name=name, index=False)

        book = xl.book
        ws = book["Index"]
        _style_header(ws, 1, index_df.shape[1]); ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions; _apply_numfmt(ws, 1, list(index_df.columns)); _autofit(ws, 1)
        for cond, name in sheet_for.items():
            ws = book[name]; ncols = ws.max_column
            headers = [ws.cell(row=1, column=j).value for j in range(1, ncols + 1)]
            _style_header(ws, 1, ncols)
            ws.freeze_panes = "C2"           # lock header + Setpoint/Severity columns
            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"
            _apply_numfmt(ws, 1, headers); _autofit(ws, 1)

    os.replace(_tmp_path, out_path)
    return {"out_path": out_path, "runs": len(folders), "conditions": len(by_cond),
            "sheets": ["Index"] + list(sheet_for.values())}
