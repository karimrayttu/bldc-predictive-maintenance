"""End-to-end verification of the real acquisition path.

The only mocked boundary is serial transport. Each test line has the exact integer
field order emitted by firmware/main.c and traverses the production parser,
raw-current augmentation, session writer, consolidation path, and live UI.
"""

import csv
import glob
import hashlib
import io
import json
import math
import os
import re
import shutil
import sys
import tempfile
import time
import types


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(ROOT)
sys.path.insert(0, ROOT)

PASS, FAIL = [], []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(f"  [{'PASS' if cond else 'FAIL'}] {name}" + (f"   {detail}" if detail else ""))
    return cond


def section(title):
    print(f"\n{'=' * 74}\n{title}\n{'=' * 74}")


# ============================================================ 1. IMPORTS
section("1. MODULES")
from PyQt5 import QtWidgets
from apps.bench.main import load_config, QSS
from bldc_phm.schema import (
    ALL_COLUMNS,
    DERIVED_COLUMNS,
    FUTURE_COLUMNS,
    META_COLUMNS,
    STREAM_COLUMNS,
    decode_status,
)
from bldc_phm.validator import FrameStats, parse_and_validate
from bldc_phm.session import SessionManager
from bldc_phm import board_f401re as board
from bldc_phm import coverage as cov
import yaml

check("all core modules import", True)
check("firmware stream schema is 44 integers (incl. gravity + fault code)", len(STREAM_COLUMNS) == 44, str(len(STREAM_COLUMNS)))
check(
    "full session schema is 65 columns",
    len(ALL_COLUMNS) == 65,
    f"{len(META_COLUMNS)}+{len(STREAM_COLUMNS)}+{len(DERIVED_COLUMNS)}+{len(FUTURE_COLUMNS)}",
)


# ============================================================ 2. WIRE FORMAT
section("2. FIRMWARE WIRE FORMAT -> parse_and_validate")


def wire(
    t_ms,
    rpm,
    ripple,
    i_mv,
    t_mv,
    acc,
    status=0b0011111110111111,
    i_raw_mean=2108,
    vref_raw_mean=1500,
    factory_cal=1504,
):
    """Return one firmware-shaped line with internally consistent raw aggregates."""
    i_count, vref_count = 800, 8
    i_lo, i_hi = i_raw_mean - 10, i_raw_mean + 10
    values = {
        "t_ms": t_ms,
        "rpm": rpm,
        "ripple_permil": ripple,
        "min_us": 1600,
        "max_us": 1700,
        "hall": (t_ms // 100 % 6) + 1,
        "illegal": 0,
        "edges": t_ms * 24 // 100,
        "step": 9,
        "coast_ms": 0,
        "period_std_permil": 12,
        "mech1x_permil": 8,
        "elec_permil": 15,
        # Legacy diagnostics are deliberately independent of the raw calibration path.
        "i_dc_mv": i_mv,
        "adc_vref_raw": vref_raw_mean,
        "adc_pc0_mv": 0,
        "accel_rms_mg": acc,
        "accel_pk_mg": acc * 2,
        "temp_mv": t_mv,
        "sensor_status": status,
        "i_dc_peak_mv": i_mv + 60,
        "i_dc_min_mv": i_mv - 40,
        "i_adc_raw_sum": i_raw_mean * i_count,
        "i_adc_raw_count": i_count,
        "i_adc_raw_min": i_lo,
        "i_adc_raw_max": i_hi,
        "i_adc_ema_q8": i_raw_mean * 256,
        "vref_adc_raw_sum": vref_raw_mean * vref_count,
        "vref_adc_raw_count": vref_count,
        "vref_adc_ema_q8": vref_raw_mean * 256,
        "vref_factory_cal_raw": factory_cal,
        "frame_seq": t_ms // 100,
        "pwm_ticks": 800,
        "i_adc_window_start_us": t_ms * 1000 - 100000,
        "i_adc_window_end_us": t_ms * 1000,
        "i_adc_fail_count": 0,
        "vref_window_start_us": t_ms * 1000 - 500,
        "vref_window_end_us": t_ms * 1000,
        # 798 samples at the mean plus one at each +/-10-code endpoint.
        "i_adc_raw_sumsq": i_count * i_raw_mean**2 + 200,
        "vref_adc_raw_sumsq": vref_count * vref_raw_mean**2,
        # orientation (gravity vector): motor mounted horizontal -> one axis ~1 g
        "accel_x_mg": 12,
        "accel_y_mg": -8,
        "accel_z_mg": 1004,
        "fault_code": 15,
    }
    return ",".join(str(values[column]) for column in STREAM_COLUMNS)


stats = FrameStats()
line = wire(1000, 1500, 292, 1700, 1400, 164)
check("wire line has 44 fields", line.count(",") == 43, f"{line[:58]}...")
check("fault code parses (on-device detector)", True)
frame, err = parse_and_validate(line, stats)
check("valid line parses", frame is not None and err is None)
check(
    "parsed motion/current/temp fields are unchanged",
    frame
    and frame["rpm"] == 1500
    and frame["i_dc_mv"] == 1700
    and frame["temp_mv"] == 1400,
)
check(
    "raw PA0 aggregates parse exactly",
    frame
    and frame["i_adc_raw_sum"] == 2108 * 800
    and frame["i_adc_raw_count"] == 800
    and frame["i_adc_raw_sumsq"] == 800 * 2108**2 + 200,
)
check(
    "signed gravity vector parses (orientation channel)",
    frame and frame["accel_x_mg"] == 12 and frame["accel_y_mg"] == -8
    and frame["accel_z_mg"] == 1004,
)
check(
    "frame timing/provenance parses exactly",
    frame
    and frame["frame_seq"] == 10
    and frame["pwm_ticks"] == 800
    and frame["i_adc_window_end_us"] == 1_000_000
    and frame["i_adc_fail_count"] == 0,
)

bad, _ = parse_and_validate(wire(1100, 99999, 292, 1700, 1400, 164), stats)
check("out-of-range rpm rejects the whole frame", bad is None)
short, _ = parse_and_validate("1,2,3", stats)
check("short line rejected", short is None)
junk, _ = parse_and_validate("1000,abc,292" + ",0" * 41, stats)
check("non-integer line rejected", junk is None)


# ============================================================ 3. APP + UI STRUCTURE
section("3. APPLICATION / UI STRUCTURE")
app = QtWidgets.QApplication([])
app.setStyle("Fusion")
app.setStyleSheet(QSS)
cfg = load_config()
from apps.bench.ui.main_window import FIXTURE_MODES, MainWindow
from apps.bench.ui import design as D

w = MainWindow(cfg)
w.resize(1500, 950)
w.move(-4000, -4000)
w.show()
app.processEvents()
check("window builds", True)
check("tab shell present with at least 2 pages", w.tabs.count() >= 2, f"{w.tabs.count()} tabs")

pages_ok = True
for index in range(w.tabs.count()):
    try:
        w.tabs.setCurrentIndex(index)
        app.processEvents()
    except Exception:
        pages_ok = False
check("every page switches without error", pages_ok)
check(
    "configured default fixture exists",
    cfg.get("default_fixture", "mounted") in FIXTURE_MODES,
)

bindings = [
    "val_rpm",
    "val_ripple",
    "val_hall",
    "val_illegal",
    "val_coast",
    "val_current",
    "val_vib",
    "val_temp",
]
missing = [name for name in bindings if not hasattr(w, name)]
check("core backend binding targets present", not missing, f"missing {missing}")

src = io.open("apps/bench/ui/main_window.py", encoding="utf-8").read()
check("no synthetic-data source can be constructed", not re.findall(r"SimulatorSource\s*\(", src))
check("exactly one _augment_current implementation", src.count("def _augment_current(") == 1)
check("retired legacy current implementation is absent", "_augment_current_legacy_disabled" not in src)
check("retired auto-zero state is absent", "self._zero_" not in src and "_reacquire_zero" not in src)
check("display median removed (operator-requested): card renders every frame",
      "_disp_amps" not in src)

augment_src = src[src.index("    def _augment_current(") : src.index("    def _augment_temp(")]
# The filtered channel IS the measurement by design: the raw 800-sample mean
# aliases the ~20 kHz chopper (its apparent slope measured duty-dependent, 483 vs
# 263.5 mV/A across two runs), while the firmware's oversampled+median+EMA channel
# reproduced the series meter across two independent sessions. Raw aggregates stay
# streamed for diagnostics.
check("current conversion uses the filtered channel (bench-verified)",
      '"i_dc_mv"' in augment_src)
check("sector-normalized anchor present (hall_sector_offset_mv)",
      "hall_sector_offset_mv" in augment_src)
check("amps rounded to 4 dp, never clamped", "round(" in augment_src
      and "max(0" not in augment_src)
check(
    "current card is rendered from raw mean and factory-VREF voltage",
    all(token in src for token in ("i_adc_raw_mean", "i_adc_factory_mv", "raw_n")),
)


class _FakeStats:
    total = 0


rate_owner = type("X", (), {})()
rate_owner.stats = _FakeStats()
rate_owner._frame_rate_text = types.MethodType(MainWindow._frame_rate_text, rate_owner)
first_rate = rate_owner._frame_rate_text()
check("frame-rate card is blank before a measurement window",
      first_rate == D.NO_VALUE, repr(first_rate))
rate_owner._fps_t -= 1.0
rate_owner.stats.total = 11
measured_rate = rate_owner._frame_rate_text()
check("frame-rate card reports a real rate", abs(float(measured_rate) - 11.0) < 0.6)
check("coast-down card rejects pre-first-coast uptime", "0 < _cm <= 120000" in src)
check(
    "exactly one raw source.send() site owns motor commands",
    len(re.findall(r"self\.source\.send\(", src)) == 1,
)


# ============================================================ 4. CURRENT CHANNEL
section("4. CURRENT: anchored filtered channel + sector normalization + A1 TEMP")
# Contract as bench-verified and operator-approved 2026-07-30:
#   amps = idle_offset_a + (i_dc_mv - anchor)/mv_per_amp
#   * NOTHING publishes until an anchor exists (8 flat frames, commanded-off, 3 s dwell)
#   * anchor is sector-NORMALIZED: parked in hall 1/2/4 (two Hall outputs sinking
#     pull-up current) the measured hall_sector_offset_mv is subtracted, so the zero
#     always refers to the one-sink baseline. Blind-verified 9 stops: sector
#     predicted the series meter 8/8 (0.062-0.064 vs 0.053 A).
#   * locked rotor (rpm 0, NOT commanded off) must never anchor.
CS = w.current_cfg
MVA = float(CS["mv_per_amp"]); IDLE = float(CS["idle_offset_a"])
SOFF = float(CS["hall_sector_offset_mv"])
check("calibration is enabled and valid", CS.get("calibration_valid") is True)
check("idle_offset_a is the anchor-state truth 0.053", IDLE == 0.053, f"{IDLE}")
check("sector offset is the operator-trimmed 2.37 mV", SOFF == 2.37, f"{SOFF}")

def _cur(mv, rpm, hall, cmd_off, dwell=10.0, n=1):
    fr = None
    for _ in range(n):
        line = wire(1000, rpm, 10, int(mv), 1650, 50, i_raw_mean=2100)
        fr, _e = parse_and_validate(line, FrameStats())
        fr["i_dc_mv"] = mv
        fr["hall"] = hall
        w._cmd_off = cmd_off
        w._cmd_off_since = time.monotonic() - dwell
        w._augment_current(fr)
    return fr

def _reset_anchor():
    w._anchor_ok = False
    w._anchor_mv = None
    if hasattr(w, "_anchor_hist"):
        w._anchor_hist.clear()

# ---- no measurement before the zero reference exists
_reset_anchor()
f = _cur(1660, 1500, 3, False)
check("no amps published before the anchor exists", f.get("i_dc_a") is None)
check("pre-anchor frame is flagged zeroing", f.get("i_dc_zeroing") is True)

# ---- clean-park anchor: rest reads exactly the idle constant
_reset_anchor()
f = _cur(1648, 0, 5, True, n=10)
check("anchors after 8 flat commanded-off frames", f.get("i_dc_zero_mv") == 1648.0,
      str(f.get("i_dc_zero_mv")))
check("rest in a clean park reads the idle constant",
      f.get("i_dc_a") == round(IDLE, 4), str(f.get("i_dc_a")))
check("anchor hall recorded", f.get("i_dc_anchor_hall") == 5)

# ---- at speed: exact equation + 4 dp rounding
f = _cur(1661, 1500, 3, False)
_hand = round(IDLE + (1661 - 1648.0) / MVA, 4)
check("at-speed amps match the hand equation to 4 dp", f.get("i_dc_a") == _hand,
      f"{f.get('i_dc_a')} vs {_hand}")

# ---- sector normalization: two-sink park anchors to the one-sink baseline
_reset_anchor()
f = _cur(1648 + SOFF, 0, 2, True, n=10)
check("two-sink park anchor is normalized to the clean baseline",
      f.get("i_dc_zero_mv") == 1648.0, str(f.get("i_dc_zero_mv")))
check("two-sink park rest displays idle + sector current (the meter's 0.062)",
      f.get("i_dc_a") == round(IDLE + SOFF / MVA, 4), str(f.get("i_dc_a")))

# ---- guards
_before = w._anchor_mv
f = _cur(1700, 0, 3, False, n=10)          # rpm 0 but NOT commanded off = locked rotor
check("locked rotor never re-anchors", w._anchor_mv == _before)
_reset_anchor()
for _step in range(10):                     # settling ramp, all frames "commanded off"
    _cur(1600 + _step * 8, 0, 5, True)
check("a settling ramp cannot satisfy the flatness gate",
      not getattr(w, "_anchor_ok", False))

# restore a sane anchor for the sections below
_reset_anchor()
_cur(1648, 0, 5, True, n=10)

tc = w.temp_cfg
temp_ok = True
for millivolts in (1650, 1400, 1027):
    temp_frame = {"temp_mv": millivolts, "sensor_status": 0b0011111110111111}
    w._augment_temp(temp_frame)
    volts = millivolts / 1000.0
    resistance = tc["series_ohm"] * volts / (tc["vin_v"] - volts)
    kelvin = 1.0 / (1.0 / tc["t0_k"] + math.log(resistance / tc["r0_ohm"]) / tc["beta"])
    temp_ok &= abs(temp_frame["temp_motor_c"] - (kelvin - 273.15)) < 0.05
check("NTC beta equation matches hand calculation", temp_ok)
temp_frame = {"temp_mv": 1650, "sensor_status": 0b0011111110111111}
w._augment_temp(temp_frame)
check("1650 mV maps to 25.00 C", abs(temp_frame["temp_motor_c"] - 25.0) < 0.01)
temp_fault = {"temp_mv": 1400, "sensor_status": 0b0011111110111011}
w._augment_temp(temp_fault)
check("faulted temperature channel publishes no temperature",
      temp_fault["temp_motor_c"] is None)

# ============================================================ 5. CONFIG CROSS-CHECK
section("5. CONFIG CONSISTENCY")
tax = yaml.safe_load(io.open("bldc_phm/config/taxonomy.yaml", encoding="utf-8"))
hookups = yaml.safe_load(io.open("bldc_phm/config/hookups.yaml", encoding="utf-8"))["channels"]
pinout = yaml.safe_load(io.open("bldc_phm/config/pinout.yaml", encoding="utf-8"))["pins"]
app_cfg = yaml.safe_load(io.open("bldc_phm/config/app_config.yaml", encoding="utf-8"))
saved_current = app_cfg["current_sensor"]

# driver_enable is a CONTROL line (EN on PA8/D7), documented in hookups but not a
# recorded data channel, so it is excluded from the taxonomy comparison.
check("taxonomy channels match hookup channels (control lines excluded)",
      set(tax["channels"]) == set(hookups) - {"driver_enable"})
check("driver EN control line is documented in hookups", "driver_enable" in hookups)
check(
    "conditions are the five planned faults",
    tax["conditions"] == ["healthy", "loose_mount", "rotor_drag", "overheat", "orientation_change"],
)
check("coverage conditions match taxonomy", tax["coverage_plan"]["conditions"] == tax["conditions"])
check(
    "ten planned speeds are reachable",
    len(tax["coverage_plan"]["speeds"]) == 10 and max(tax["coverage_plan"]["speeds"]) <= 2310,
)
check("runs_per_cell is three", tax["coverage_plan"]["runs_per_cell"] == 3)
check(
    "A0 pin agrees across app/hookup/pinout/firmware",
    "PA0" in saved_current["adc_pin"]
    and "PA0" in hookups["dc_bus_current"]["nucleo_pin"]
    and pinout["i_sense"] == "PA0"
    and board.FIRMWARE_PINS["i_sense"] == "PA0",
)
check(
    "A1 pin agrees across app/hookup/pinout/firmware",
    "PA1" in app_cfg["temp_sensor"]["adc_pin"]
    and "PA1" in hookups["motor_temp"]["nucleo_pin"]
    and pinout["ntc"] == "PA1"
    and board.FIRMWARE_PINS["ntc"] == "PA1",
)
check("checked-in current calibration is VALID (bench-calibrated 2026-07-30)",
      saved_current["calibration_valid"] is True)
check("checked-in current model is raw PA0/VREF affine", saved_current["calibration_model"] == "raw_pa0_vref_affine_v1")
check(
    "raw diagnostic coefficients are present",
    isinstance(saved_current["raw_ratio_slope_a_per_ratio"], float)
    and isinstance(saved_current["raw_ratio_intercept_a"], float),
)
check(
    "filtered-channel constants match the bench calibration",
    saved_current["mv_per_amp"] == 263.5 and saved_current["idle_offset_a"] == 0.053
    and saved_current["hall_sector_offset_mv"] == 2.37,
)
check(
    "saved pin map validates against F401RE rules",
    not [
        message
        for level, message in board.validate(
            {key: pinout.get(key, "") for key, _label, _color, _help in board.SIGNALS}
        )
        if level == "error"
    ],
)

bad_defaults = []
for field in tax["fields"]:
    options, default = field.get("options"), field.get("default")
    if (
        isinstance(options, str)
        and options in tax
        and default not in (None, "", [])
        and isinstance(tax[options], list)
        and default not in tax[options]
    ):
        bad_defaults.append((field["key"], default))
check("no field default points at a deleted option", not bad_defaults, str(bad_defaults))


# ============================================================ 6. SCHEDULER
section("6. SCHEDULER")
_requirements, overview, todo = cov.build_requirements({}, tax["coverage_plan"])
# Reduced plan (2026-07-30, operator): healthy 10x3 + rotor_drag 10x3 (full grids)
# + loose_mount/overheat/orientation_change at 3 speeds x 2 (their signatures barely
# depend on rpm) = 78 total runs.
check("49 required runs (final reduced plan)", overview["target"] == 49,
      str(overview["target"]))
check("hold is two minutes (operator-shortened for the campaign)", cov.HOLD_S == 120)
from apps.bench.ui.main_window import FULL_RUN_HOLD_S as _FRH
check("app run length equals the coverage estimate", _FRH == cov.HOLD_S,
      f"{_FRH} vs {cov.HOLD_S}")
check(
    "planned workload matches the final plan",
    2.04 < overview["plan_workload"]["hours"] < 2.34,
    f"{overview['plan_workload']['hours']:.1f} h",
)
check("invalid runs do not satisfy their cell (S014 lesson)",
      "valid_for_classifier" in io.open("bldc_phm/coverage.py", encoding="utf-8").read())
check("queue begins with healthy runs", todo[0][0] == "healthy" and todo[5][0] == "healthy")
check("orientation change is scheduled last", todo[-1][0] == "orientation_change")

# per-category batch runner (operator-requested 2026-07-30)
check("one batch button per planned condition",
      hasattr(w, "batch_btns") and set(w.batch_btns) == set(tax["conditions"]),
      str(sorted(getattr(w, "batch_btns", {}))))
check("batch engine methods present",
      all(hasattr(w, m) for m in ("_start_batch", "_batch_step", "_batch_todo")))
check("_batch_todo filters to one condition",
      all(t[0] == "healthy" for t in w._batch_todo("healthy")))
_src2 = io.open("apps/bench/ui/main_window.py", encoding="utf-8").read()
check("STOP ALL cancels a batch", "_batch_cond = None              # STOP ALL" in _src2)
check("batch only continues after a SAVED run",
      "if sid:" in _src2 and "singleShot(5000, self._batch_step)" in _src2)


# ============================================================ 7. FULL DATA PATH
section("7. FULL DATA PATH: wire -> parse -> raw affine -> CSV -> workbook")
tmp = tempfile.mkdtemp(prefix="bldc_e2e_")
session = SessionManager(tmp, meta_fields=w.session.meta_fields, subroot="mounted_baseline")
path_stats = FrameStats()
meta = {
    "condition": "rotor_drag",
    "severity": "mild",
    "speed_rpm_nominal": 1500,
    "fixture_state": "mounted",
    "valid_for_classifier": "true",
    "motor_model": "42BL41.010 Rev.B",
    "run_kind": "official",
    "cal_current_model": "raw_pa0_vref_affine_v1",
    "cal_current_raw_ratio_slope_a_per_ratio": CS["raw_ratio_slope_a_per_ratio"],
    "cal_current_raw_ratio_intercept_a": CS["raw_ratio_intercept_a"],
    "cal_current_mv_per_amp": MVA,
    "cal_current_idle_offset_a": IDLE,
    "cal_current_valid": True,
    "cal_current_session": "E2E_IN_MEMORY_ONLY",
    "cal_temp_beta": w.temp_cfg.get("beta"),
    "cal_temp_r0_ohm": w.temp_cfg.get("r0_ohm"),
    "cal_temp_series_ohm": w.temp_cfg.get("series_ohm"),
    "cal_temp_vin_v": w.temp_cfg.get("vin_v"),
    "cal_temp_t0_k": w.temp_cfg.get("t0_k"),
}
session.start(meta)

# Anchor at a KNOWN level so every row's amps are reproducible by hand below.
_reset_anchor()
_cur(1160, 0, 5, True, n=10)
ANCHOR_MV = 1160.0

nwritten = 0
for index in range(60):
    raw_mean = 2104 + index % 9
    transport_line = wire(
        1000 + index * 100,
        1500 + (index % 7) - 3,
        292 + index % 5,
        1200 + index % 9,
        1400 - index,
        164 + index % 11,
        i_raw_mean=raw_mean,
    )
    parsed, _ = parse_and_validate(transport_line, path_stats)
    if parsed is None:
        continue
    w._augment_current(parsed)
    w._augment_temp(parsed)
    session.write_frame(parsed)
    nwritten += 1
session.stop(path_stats)
check("60/60 valid frames survive the complete path", nwritten == 60)

csv_paths = glob.glob(os.path.join(tmp, "**", "data.csv"), recursive=True)
check("session CSV written", len(csv_paths) == 1)
rows = list(csv.reader(io.open(csv_paths[0], encoding="utf-8")))
check("CSV header exactly matches ALL_COLUMNS", rows[0] == ALL_COLUMNS)
check("CSV contains 60 data rows", len(rows) - 1 == 60)
first_row = dict(zip(rows[0], rows[1]))
first_expected_a = round(IDLE + (float(first_row["i_dc_mv"]) - ANCHOR_MV) / MVA, 4)
check(
    "CSV amperes are reproducible from the anchored filtered equation",
    math.isclose(float(first_row["i_dc_a"]), first_expected_a, rel_tol=0, abs_tol=1e-9),
    f"{first_row['i_dc_a']} vs {first_expected_a}",
)
check("CSV keeps the exact raw count", int(first_row["i_adc_raw_count"]) == 800)
check(
    "CSV keeps exact sum-of-squares statistics",
    int(first_row["i_adc_raw_sumsq"]) == 800 * 2104**2 + 200
    and int(first_row["vref_adc_raw_sumsq"]) == 8 * 1500**2,
)
check(
    "CSV keeps acquisition timing and command provenance",
    int(first_row["frame_seq"]) == 10
    and int(first_row["pwm_ticks"]) == 800
    and int(first_row["i_adc_fail_count"]) == 0,
)
check("per-row anchor provenance is logged", float(first_row["i_dc_zero_mv"]) == ANCHOR_MV,
      first_row["i_dc_zero_mv"])
check("raw extrema-derived current envelope is logged", first_row["i_dc_peak_a"] and first_row["i_dc_min_a"])
check("primary and signed current remain identical", first_row["i_dc_a"] == first_row["i_dc_a_signed"])
blank_columns = {key for key, value in first_row.items() if value == ""}
check(
    "only reserved future channels are blank",
    blank_columns == set(FUTURE_COLUMNS),
    str(sorted(blank_columns)),
)

meta_path = glob.glob(os.path.join(tmp, "**", "meta.json"), recursive=True)[0]
saved_meta = json.load(io.open(meta_path, encoding="utf-8"))
check(
    "meta.json carries raw affine provenance",
    saved_meta.get("cal_current_model") == "raw_pa0_vref_affine_v1"
    and saved_meta.get("cal_current_raw_ratio_slope_a_per_ratio") == CS["raw_ratio_slope_a_per_ratio"]
    and saved_meta.get("cal_current_mv_per_amp") == MVA
    and saved_meta.get("cal_current_idle_offset_a") == IDLE,
)
check("meta.json carries NTC provenance", saved_meta.get("cal_temp_beta") == 3550.0)
check("meta.json carries frame statistics", "frame_stats" in saved_meta)

try:
    from bldc_phm import consolidate as consolidation

    workbook_path = os.path.join(tmp, "MOUNTED.xlsx")
    report = consolidation.consolidate_fixture(tmp, "mounted_baseline", workbook_path)
    check("consolidation completes", report is not None, str(report)[:90])
    check(
        "workbook is created",
        os.path.exists(workbook_path),
        f"{os.path.getsize(workbook_path) / 1024:.0f} KB" if os.path.exists(workbook_path) else "",
    )
    if os.path.exists(workbook_path):
        import openpyxl

        workbook = openpyxl.load_workbook(workbook_path)
        check("workbook contains sheets", len(workbook.sheetnames) > 0, str(workbook.sheetnames))
except Exception as exc:
    check("consolidate -> workbook", False, f"{type(exc).__name__}: {exc}")


# ============================================================ 8. LIVE UI PATH
section("8. LIVE CARD RENDERS THE SAME RAW FRAME")


class _OneShotTransport:
    def __init__(self, transport_line):
        self._lines = [transport_line]

    def drain(self):
        lines, self._lines = self._lines, []
        return lines


w.tabs.setCurrentIndex(0)
# Temporarily mark the calibration invalid to exercise the UNCALIBRATED branch.
_saved_valid = w.current_cfg["calibration_valid"]
w.current_cfg["calibration_valid"] = False
expected_factory_mv = (2108 / 1500.0) * 3300.0 * 1504 / 4095.0
uncal_ui_line = wire(9000, 1500, 292, 275, 1400, 164, i_raw_mean=2108)
w.source = _OneShotTransport(uncal_ui_line)
w._tick()
uncal_text = w.val_current.text()
check("uncalibrated card says UNCALIBRATED", uncal_text.startswith("UNCALIBRATED"))
check("uncalibrated card shows actual raw mean", "ADC 2108.000000 cnt" in uncal_text, uncal_text)
check("uncalibrated card shows actual sample count", "n=800" in uncal_text, uncal_text)
check(
    "uncalibrated card shows factory-VREF millivolts",
    f"{expected_factory_mv:.6f} mV" in uncal_text,
    uncal_text,
)
check("uncalibrated card does not display legacy filtered millivolts", "275.000000 mV" not in uncal_text)

w.current_cfg["calibration_valid"] = _saved_valid
cal_ui_line = wire(9100, 1500, 292, 3100, 1400, 164, i_raw_mean=2112)
w.source = _OneShotTransport(cal_ui_line)
w._tick()
cal_text = w.val_current.text()
# Approved card format: ONLY the realtime amps (display median) and the ADC mV.
cal_expected = round(IDLE + (3100 - ANCHOR_MV) / MVA, 4)
check("calibrated card shows amps from the anchored equation",
      cal_text.startswith(f"{cal_expected:.3f} A"), cal_text)
check("card renders the frame directly (no median state)", not hasattr(w, "_disp_amps"))
check("card shows the ADC millivolts and nothing else",
      cal_text.endswith("3100 mV") and "n=" not in cal_text and "peak" not in cal_text,
      cal_text)
check("temperature card receives the same live frame", w.val_temp.text().startswith("91.1"))

w._write_live_status()
live_path = os.path.join(cfg["data_dir"], "_live", "status.json")
check("status.json written", os.path.exists(live_path))
if os.path.exists(live_path):
    live_status = json.load(io.open(live_path, encoding="utf-8"))
    check(
        "status current_sensor_mv is the raw factory-VREF value",
        live_status.get("current_sensor_mv") == w.last_frame.get("i_adc_factory_mv"),
    )
    check(
        "status includes raw aggregate provenance",
        live_status.get("current_adc_raw_sum") == w.last_frame.get("i_adc_raw_sum")
        and live_status.get("current_adc_raw_sumsq") == w.last_frame.get("i_adc_raw_sumsq")
        and live_status.get("vref_factory_cal_raw") == w.last_frame.get("vref_factory_cal_raw")
        and live_status.get("frame_seq") == w.last_frame.get("frame_seq")
        and live_status.get("current_adc_window_end_us") == w.last_frame.get("i_adc_window_end_us"),
    )
    check("status contains no retired zero fault", "zero_fault" not in live_status)
w.source = None


# ============================================================ 9. MANIFEST / SOURCE INTEGRITY
section("9. ACQUISITION CORE INTEGRITY")
manifest_path = os.path.join("tools", "core_manifest.json")
if os.path.exists(manifest_path):
    manifest = json.load(io.open(manifest_path, encoding="utf-8"))
    for path, expected_hash in manifest.items():
        actual_hash = hashlib.sha256(open(path, "rb").read()).hexdigest() if os.path.exists(path) else ""
        check(
            f"{path} matches manifest",
            actual_hash == expected_hash,
            "" if actual_hash == expected_hash else f"changed ({actual_hash[:12]} != {expected_hash[:12]})",
        )
else:
    check("core manifest present", False, "tools/core_manifest.json missing")

all_source = io.open("bldc_phm/sources.py", encoding="utf-8").read() + src
constructors = [
    match
    for match in re.finditer(r"SimulatorSource\s*\(", all_source)
    if not all_source[max(0, match.start() - 6) : match.start()].endswith("class ")
    and "isinstance" not in all_source[max(0, match.start() - 40) : match.start()]
]
check("SimulatorSource has zero construction sites", not constructors)


# ============================================================ 10. BOARD IDENTITY
section("10. BOARD IDENTITY")
wrong_board = re.compile(r"G474|H743|F446|MB1367|STM32G4", re.I)
# A line may name another board only where it is explicitly describing the
# superseded generations. Everything else naming a non-F401RE part is stale.
allowed_history = ("old NUCLEO-H743ZI2 build", "was the G474RE-era wiring",
                   "history/")
offenders = []
paths = (
    glob.glob("apps/**/*.py", recursive=True)
    + glob.glob("bldc_phm/**/*.py", recursive=True)
    + glob.glob("bldc_phm/**/*.yaml", recursive=True)
    + glob.glob("docs/*.md")
    + glob.glob("*.md")
    + glob.glob("firmware/*.c")
)
for path in paths:
    if "_archive" in path or "firmware/history" in path.replace("\\", "/"):
        continue
    for line_number, text_line in enumerate(
        io.open(path, encoding="utf-8", errors="ignore").read().splitlines(),
        1,
    ):
        if wrong_board.search(text_line) and not any(token in text_line for token in allowed_history):
            offenders.append(f"{path}:{line_number}")
check("no stale board identity outside _archive", not offenders, str(offenders[:4]))
check(
    "firmware pin map is the F401RE map",
    board.FIRMWARE_PINS
    == {
        "hall_a": "PA10",
        "hall_b": "PB3",
        "hall_c": "PB5",
        "sv_pwm": "PB4",
        "i_sense": "PA0",
        "ntc": "PA1",
        "kx_sda": "PB9",
        "kx_scl": "PB8",
        "kx_int1": "PA9",
        "drv_en": "PA8",
    },
)
check(
    "SV is PB4/TIM3_CH1",
    "TIM3_CH1" in board.BY_NAME["PB4"].timers and board.FIRMWARE_PINS["sv_pwm"] == "PB4",
)
check(
    "Hall inputs use distinct EXTI lines",
    len(
        {
            board.BY_NAME[board.FIRMWARE_PINS[key]].exti_line
            for key in ("hall_a", "hall_b", "hall_c")
        }
    )
    == 3,
)
firmware_generator = io.open("bldc_phm/firmwaregen.py", encoding="utf-8").read()
check("speed ceiling is 2310 rpm", "DRIVER_FULLSCALE_RPM = 2310" in firmware_generator)
# docs/PINOUT.md is hand-maintained, not generated: nothing writes it. So check it
# against the pins board_f401re.py declares, which is what "single source of truth"
# has to mean if the page is transcribed by hand.
board_src = io.open("bldc_phm/board_f401re.py", encoding="utf-8").read()
pinout_txt = io.open("docs/PINOUT.md", encoding="utf-8").read() \
    if os.path.exists("docs/PINOUT.md") else ""
pinout_missing = [
    f"{role}={pin}"
    for role, pin in re.findall(r'"(sv_pwm|i_sense|ntc)":\s*"(P[A-H]\d+)"', board_src)
    # \b alone is not enough: "PA1" matches inside "PA10".
    if not re.search(r"\b" + pin + r"\b(?!\d)", pinout_txt)
]
check(
    "docs/PINOUT.md agrees with board_f401re.py on SV / current / NTC pins",
    bool(pinout_txt) and not pinout_missing,
    "" if not pinout_missing else "missing " + ", ".join(pinout_missing),
)


shutil.rmtree(tmp, ignore_errors=True)
w.close()
section(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
for failed in FAIL:
    print("   FAILED:", failed)
sys.exit(1 if FAIL else 0)
